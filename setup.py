import os
import json
from sqlalchemy.orm import Session
from models import get_db, User, MenuItem
from crud import create_user, create_menu_item
from auth import get_password_hash

SETUP_FILE = "setup_complete.json"

def is_setup_complete():
    return os.path.exists(SETUP_FILE)

def save_setup_config(config):
    # Remove binary data before saving to JSON
    config_to_save = config.copy()
    if 'menu_file' in config_to_save:
        del config_to_save['menu_file']  # Remove binary data
    if 'menu_filename' in config_to_save:
        del config_to_save['menu_filename']  # Remove filename too
    
    with open(SETUP_FILE, 'w') as f:
        json.dump(config_to_save, f)

def get_setup_config():
    if os.path.exists(SETUP_FILE):
        try:
            with open(SETUP_FILE, 'r') as f:
                return json.load(f)
        except (json.JSONDecodeError, IOError) as e:
            print(f"Warning: Corrupted setup file, removing: {e}")
            os.remove(SETUP_FILE)
    return {}

def apply_setup(config):
    db = next(get_db())
    
    try:
        # Create admin user
        print(f"Creating admin user: {config['admin_username']}")
        create_user(db, config['admin_username'], config['admin_password'], 'admin')
        
        # Process menu file if provided
        if config.get('menu_file'):
            if isinstance(config['menu_file'], bytes):
                # Handle uploaded file content
                process_uploaded_menu(db, config['menu_file'], config.get('menu_filename', 'menu.xlsx'))
            else:
                # Handle file path
                process_menu_file(db, config['menu_file'])
        
        # Save configuration
        save_setup_config(config)
        
        print("Setup completed successfully!")
        print(f"Restaurant: {config['restaurant_name']}")
        print(f"Admin: {config['admin_username']}")
        
    except Exception as e:
        print(f"Setup failed: {e}")
        raise
    finally:
        db.close()

def process_uploaded_menu(db: Session, file_content: bytes, filename: str):
    if filename.endswith(('.xlsx', '.xls')):
        process_excel_content(db, file_content)
    elif filename.endswith('.pdf'):
        process_pdf_content(db, file_content)

def process_excel_content(db: Session, file_content: bytes, restaurant_id: int = None):
    try:
        from tenant import get_current_restaurant_id
        from models import MenuItem
        if restaurant_id is None:
            try:
                restaurant_id = get_current_restaurant_id()
            except:
                restaurant_id = 1  # Default to first restaurant
        
        # Clear existing menu items for this restaurant only
        deleted_count = db.query(MenuItem).filter(MenuItem.restaurant_id == restaurant_id).delete()
        db.commit()
        print(f"Cleared {deleted_count} existing menu items for restaurant {restaurant_id}")
        
        import openpyxl
        import io
        wb = openpyxl.load_workbook(io.BytesIO(file_content))
        ws = wb.active
        
        items_added = 0
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if row and len(row) >= 3 and row[0] and row[1] and row[2]:
                name = str(row[0]).strip()
                ingredients = str(row[1]).strip()
                price = float(row[2])
                category = str(row[3]).strip() if len(row) > 3 and row[3] else 'Food'
                
                create_menu_item(db, name, ingredients, price, category, restaurant_id)
                items_added += 1
        
        print(f"Replaced menu with {items_added} items for restaurant {restaurant_id}")
        
    except Exception as e:
        print(f"Error processing Excel file: {e}")

def process_pdf_content(db: Session, file_content: bytes, restaurant_id: int = None):
    try:
        from tenant import get_current_restaurant_id
        from models import MenuItem
        if restaurant_id is None:
            try:
                restaurant_id = get_current_restaurant_id()
            except:
                restaurant_id = 1  # Default to first restaurant
        
        # Clear existing menu items for this restaurant only
        deleted_count = db.query(MenuItem).filter(MenuItem.restaurant_id == restaurant_id).delete()
        db.commit()
        print(f"Cleared {deleted_count} existing menu items for restaurant {restaurant_id}")
        
        import PyPDF2
        import io
        pdf_reader = PyPDF2.PdfReader(io.BytesIO(file_content))
        text = ""
        for page in pdf_reader.pages:
            text += page.extract_text()
        
        # Simple parsing - expects format: Name - Price - Ingredients - Category
        import re
        pattern = r'(.+?)\s*-\s*\$?(\d+\.?\d*)\s*-\s*(.+?)\s*-\s*(.+?)(?=\n|$)'
        matches = re.findall(pattern, text, re.MULTILINE)
        
        items_added = 0
        for match in matches:
            if len(match) >= 4:
                name, price, ingredients, category = match
                name = name.strip()
                price = float(price)
                ingredients = ingredients.strip()
                category = category.strip()
                
                create_menu_item(db, name, ingredients, price, category, restaurant_id)
                items_added += 1
        
        print(f"Replaced menu with {items_added} items for restaurant {restaurant_id}")
        
    except Exception as e:
        print(f"Error processing PDF file: {e}")

def process_menu_file(db: Session, file_path: str):
    if not os.path.exists(file_path):
        return
    
    if file_path.endswith(('.xlsx', '.xls')):
        process_excel_menu(db, file_path)
    elif file_path.endswith('.pdf'):
        process_pdf_menu(db, file_path)

def process_excel_menu(db: Session, file_path: str):
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        items_added = 0
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if row and len(row) >= 3 and row[0] and row[1] and row[2]:
                name = str(row[0]).strip()
                ingredients = str(row[1]).strip()
                price = float(row[2])
                category = str(row[3]).strip() if len(row) > 3 and row[3] else 'Food'
                
                create_menu_item(db, name, ingredients, price, category)
                items_added += 1
        
        print(f"Added {items_added} menu items from Excel")
        
    except ImportError:
        print("Excel support not available. Install openpyxl: pip install openpyxl")
    except Exception as e:
        print(f"Error processing Excel file: {e}")

def process_pdf_menu(db: Session, file_path: str):
    try:
        import PyPDF2
        with open(file_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            text = ""
            for page in pdf_reader.pages:
                text += page.extract_text()
        
        # Simple parsing - expects format: Name - Price - Ingredients - Category
        import re
        pattern = r'(.+?)\s*-\s*\$?(\d+\.?\d*)\s*-\s*(.+?)\s*-\s*(.+?)(?=\n|$)'
        matches = re.findall(pattern, text, re.MULTILINE)
        
        items_added = 0
        for match in matches:
            if len(match) >= 4:
                name, price, ingredients, category = match
                name = name.strip()
                price = float(price)
                ingredients = ingredients.strip()
                category = category.strip()
                
                create_menu_item(db, name, ingredients, price, category)
                items_added += 1
        
        print(f"Added {items_added} menu items from PDF")
        
    except Exception as e:
        print(f"Error processing PDF file: {e}")